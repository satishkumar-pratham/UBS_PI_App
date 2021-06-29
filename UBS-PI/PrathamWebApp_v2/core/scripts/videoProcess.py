"""

This file is created as template for all driver codes, this will cater to transcript generation, processing, audio processing, video processing and report generation.

The basic structure is as follows :

1) read rows from database where respective flag is not set (for example for sentiment : where isTranscriptProcessed = False should be retrieved)
2) iterate over each row and perform required operation
3) for each row set the flag to true EVEN IF ERROR OCCURS
4) Save the modified row so that updates are made in the database
5) Generate logs of error if required
6) this complete code must be part of a 'run' function
7) Also Use the THRESHOLD table to identify required value


"""

#COMMENT IN PROD
import warnings
warnings.filterwarnings("ignore")


from core.models import INTERVIEW
from core.models import THRESHOLD
import os
'''
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, colors,Alignment
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
'''
from datetime import datetime
import dlib
import cv2
import math
import numpy as np
import pandas as pd
from fnmatch import fnmatch
from core.scripts.detectors import *
import traceback

shapePredictorPath = "core/scripts/shape_predictor_68_face_landmarks.dat"

def checkEyeContact(shape,frame,rgb):
    eyeContact = "false"
    #Left Eye
    ly1,ly2,lx1,lx2=extract_left_eye(shape,frame,rgb)
    lefteye=frame[ly1:ly2,lx1:lx2]
    lefteye=cv2.adaptiveThreshold(lefteye,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY,11,3)
    lefteye_contour=return_biggest_contour_in_image(lefteye)

    #Right Eye
    ry1,ry2,rx1,rx2=extract_right_eye(shape,frame,rgb)
    righteye=frame[ry1:ry2,rx1:rx2]
    righteye=cv2.adaptiveThreshold(righteye,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,cv2.THRESH_BINARY,11,2)
    righteye_contour=return_biggest_contour_in_image(righteye)

    if righteye_contour is None:
        return eyeContact

    #calibrate eyes for finding the center point
    left_v,left_h=detect_contour_location(lefteye_contour,cv2)
    right_v,right_h=detect_contour_location(righteye_contour,cv2)
    return left_v,left_h,right_v,right_h


def return_biggest_contour_in_image(frame):
    contours,_=cv2.findContours(frame,cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    contours=sorted(contours,key=lambda x:cv2.contourArea(x),reverse=True)
    #print(len(contours))
    if len(contours)>0:
        return contours[0]
    else:
        return None


def extract_left_eye(shape_predictor,frame,rgb):
    x1_left=shape_predictor.part(36).x
    x2_left=shape_predictor.part(39).x
    y1_left=shape_predictor.part(37).y
    y2_left=shape_predictor.part(40).y
    #cv2.rectangle(rgb, (x1_left-5, y1_left-5), (x2_left+5, y2_left+5), (0, 0, 189), 2)
    return y1_left-5,y2_left+5,x1_left-5,x2_left+5

def extract_right_eye(shape_predictor, frame,rgb):
    x1_right=shape_predictor.part(42).x
    x2_right=shape_predictor.part(45).x
    y1_right=shape_predictor.part(43).y
    y2_right=shape_predictor.part(46).y
    #cv2.rectangle(rgb, (x1_right-5, y1_right-5), (x2_right+5, y2_right+5), (0, 0, 189), 2)
    return y1_right-5,y2_right+5,x1_right-5,x2_right+5


## this function find the midpoint of the cornea
def detect_contour_location(frame,cv2):
    (x,y,w,h)=cv2.boundingRect(frame)
    return float(x)+float(w/2), float(y)+float(h/2)


def detect_eye_contact(vertical_line,horizontal_line,ref_vertical_line,ref_horizontal_line):
    vertical_line_absolute=abs(vertical_line-ref_vertical_line)
    horizontal_line_absolute=abs(horizontal_line-ref_horizontal_line)
    print(vertical_line," ",horizontal_line,ref_vertical_line,ref_horizontal_line)
    if vertical_line_absolute>=2.5 or horizontal_line_absolute>=2.5:
        return False
    else:
        return True


def increase_contrast(frame,cv2):
    lab= cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    cl = clahe.apply(l)
    limg = cv2.merge((cl,a,b))
    final = cv2.cvtColor(limg, cv2.COLOR_LAB2BGR)

    return final

def determine_fps_of_video(cap,cv2):
    fps = cap.get(cv2.CAP_PROP_FPS)
    print("FPS is: ",fps)
    return fps


def ResizeWithAspectRatio(image, width=None, height=None, inter=cv2.INTER_AREA):
    dim = None
    (h, w) = image.shape[:2]

    if width is None and height is None:
        return image
    if width is None:
        r = height / float(h)
        dim = (int(w * r), height)
    else:
        r = width / float(w)
        dim = (width, int(h * r))

    return cv2.resize(image, dim, interpolation=inter)

def adjust_gamma(image, gamma=1.5):
	invGamma = 1.0 / gamma
	table = np.array([((i / 255.0) ** invGamma) * 255
		for i in np.arange(0, 256)]).astype("uint8")
	return cv2.LUT(image, table)

def extract_smile(shape,frame):
    x1=shape.part(48).x
    x2=shape.part(54).x
    y1=shape.part(50).y
    y2=shape.part(57).y
    horizontal_offset = round(0.09 * (x2-x1))
    vertical_offset = round(0.06 * (y2-y1))
    return x1-horizontal_offset,y1-vertical_offset,(x2-x1+2*horizontal_offset),(y2-y1+2*vertical_offset)

def calculate_distance(shape,a,b):
    return math.sqrt((shape.part(a).x-shape.part(b).x)**2+(shape.part(a).y-shape.part(b).y)**2)

def classify_the_smile(shape,frame):
    L1=calculate_distance(shape,50,58)
    L2=calculate_distance(shape,51,57)
    L3=calculate_distance(shape,52,56)
    D1=calculate_distance(shape,48,54)
    DM=calculate_distance(shape,0,16)
    D=D1*100/DM
    if L1+L2+L3 == 0:
        return 1000,D
    return D/(L1+L2+L3),D

def findLine(shape,p1,p2):
    x1 = shape.part(p1).x
    y1 = shape.part(p1).y
    x2 = shape.part(p2).x
    y2 = shape.part(p2).y
    m = (y2-y1)/(x2-x1)
    c = y2 - m*x2
    return m,c

def faceTouch(HandsDetector,FaceDetector,rgb):
    hands = HandsDetector.detect(rgb)
    face = FaceDetector.detect(rgb)
    faceTouch = "false"
    if objects_touch(face, hands):
        img_detected = add_objects_to_image(rgb, hands)
        img_detected = add_objects_to_image(img_detected, face)
        faceTouch = "true"
    return faceTouch

def findLipBoundaryDistance(shape):
    m,c1 = findLine(shape,50,52)
    y1 = shape.part(48).y
    x1 = shape.part(48).x
    y2 = shape.part(54).y
    x2 = shape.part(54).x
    dc1 = y1-m*x1
    dc2 = y2-m*x2
    D1 = abs(c1-dc1)/math.sqrt(1+(m*m))
    D2 = abs(c1-dc2)/math.sqrt(1+(m*m))
    return min(D1,D2)


def findSmileDiameter(shape):
    D1=calculate_distance(shape,48,54)
    DM=calculate_distance(shape,0,16)
    AD=D1*100/DM
    return AD

def pointDistance(x1,y1,x2,y2):
    ans = (x2-x1)*(x2-x1)+(y2-y1)*(y2-y1)
    return math.sqrt(ans)

def findContour(shape):
    j1 = pointDistance(shape.part(0).x, shape.part(0).y, shape.part(16).x,shape.part(16).y)
    j2 = pointDistance(shape.part(1).x, shape.part(1).y, shape.part(15).x,shape.part(15).y)
    j3 = pointDistance(shape.part(2).x, shape.part(2).y, shape.part(14).x,shape.part(14).y)
    j4 = pointDistance(shape.part(3).x, shape.part(3).y, shape.part(13).x,shape.part(13).y)
    j5 = pointDistance(shape.part(6).x, shape.part(6).y, shape.part(10).x,shape.part(10).y)
    return round(j1,2),round(j2,2),round(j3,2),round(j4,2),round(j5,2)


def faceIndex(row,thresh):
    if row['Sum']<thresh:
        return 1
    else:
        return 1

def interviewAnalysis(path):

    global shapePredictorPath

    distance_smile = 0
    angle_smile = 0
    touch_lips = 0
    total_frames_count = 0
    eyeContact_frames_count = 0
    eyes_calibrated = False
    angleThreshold = 8
    distanceThreshold = 0
    smileDiameterThreshold = 0
    HandsDetector = TSDetector()
    FaceDetector = CVLibDetector()
    faceTouched = 0
    lefteye_vertical_line = 0
    lefteye_horizontal_line = 0
    righteye_vertical_line = 0
    righteye_horizontal_line = 0
    df = pd.DataFrame(columns = ['Frame', 'DistanceValue', 'AngleValue', 'SmileDiameter','Sum','Left_V','Left_H','Right_V','Right_H','FaceTouch'])
    cmin = 1000000
    cmax = 0
    detector=dlib.get_frontal_face_detector()
    predictor=dlib.shape_predictor(shapePredictorPath)
    cap=cv2.VideoCapture(path)
    frame_count = 0
    fps = determine_fps_of_video(cap,cv2)//2
    while True:
        if total_frames_count%1==0:
            ret, cframe = cap.read()
            frame_count = frame_count+1
            print("Frame Count: ",frame_count," Frame Parsed: ",total_frames_count)
            if ret and cframe is not None:
                #frame_count = frame_count + 1
                if frame_count<total_frames_count*fps:
                    continue
                cframe = ResizeWithAspectRatio(cframe, width=600)
                rgb = cv2.cvtColor(cframe, cv2.COLOR_BGR2RGB)
                frame = cv2.cvtColor(cframe,cv2.COLOR_BGR2GRAY)
                faces = detector(frame)
                if len(faces) == 0:
                    continue
                shape = predictor(frame,faces[0])
                left_v,left_h,right_v,right_h = checkEyeContact(shape,frame,cframe)
                j1,j2,j3,j4,j5 = findContour(shape)
                contourSum = round(j1+j2+j3+j4+j5,2)
                cmin = min(contourSum,cmin)
                cmax = max(contourSum,cmax)
                smile = extract_smile(shape,frame)
                isFaceTouch = faceTouch(HandsDetector, FaceDetector, rgb)
                if smile is None:
                        continue
                total_frames_count=total_frames_count+1
                MAR,D=classify_the_smile(shape,frame)
                SV=findLipBoundaryDistance(shape)
                df = df.append({'Frame' : total_frames_count, 'DistanceValue' : MAR, 'AngleValue' : SV, 'SmileDiameter': D,'Sum':contourSum,'Left_V': left_v,'Left_H': left_h,'Right_V': right_v,'Right_H': right_h,'FaceTouch': isFaceTouch},
                ignore_index = True)
                #cv2.imshow('Original Image',cframe)
                print(total_frames_count,left_v,left_h,right_v,right_h)

                if cv2.waitKey(1000) & 0xFF == ord('q'):
                    cap.release()
                    cv2.destroyAllWindows()
                    break
            else:
                print("Exiting...")
                cap.release()
                cv2.destroyAllWindows()
                break
        else:
            print("Exiting...")
            cap.release()
            cv2.destroyAllWindows()
            break
    print("Video Parsed!!")
    total_index = len(df.index)
    if total_index == 0:
        return "Error","Failed to Detect Face","Failed to Detect Face","Failed to Detect Face","Failed to Detect Face"
    print("Total Frames Detected: ",total_index)
    cap = min(total_index,50)
    mid = (cmax+cmin)/2
    df['face_label'] = df.apply(lambda row: faceIndex(row,mid), axis=1)
    print("Face Label Added")
    countLabel = []
    countLabel.append(df[df.face_label == 1].count()["Frame"])
    countLabel.append(df[df.face_label == 2].count()["Frame"])
    count = 0
    findLabel = 1
    if countLabel[0] < countLabel[1]:
        findLabel = 2
    eyeCalibrated = False
    ind = 0
    while eyeCalibrated is False:
        if df.loc[ind].face_label == findLabel:
            left_cal_h = df.loc[ind].Left_H
            left_cal_v = df.loc[ind].Left_V
            right_cal_h = df.loc[ind].Right_H
            right_cal_v = df.loc[ind].Right_V
            eyeCalibrated = True
        ind = ind + 1
    print("Threshold Frame Count: ",cap)
    while count < cap:
        ind = np.random.randint(total_index,size=1)
        ind = ind[0]
        print(count)
        if df.loc[ind].face_label == findLabel:
            distanceThreshold = distanceThreshold + df.loc[ind].DistanceValue/cap
            #angleThreshold = angleThreshold + df.loc[ind].AngleValue/50
            smileDiameterThreshold = smileDiameterThreshold + df.loc[ind].SmileDiameter/cap
            count = count+1
    distanceThreshold = 1.2*distanceThreshold
    smileDiameterThreshold = 1.1*smileDiameterThreshold
    angleThreshold = 0.8*angleThreshold
    touch_lip_threshold = 0.9*angleThreshold
    print("Distance Threshold: ",distanceThreshold)
    print("Angle Threshold: ",angleThreshold)
    print("Smile Threshold: ",smileDiameterThreshold)
    print("Touch Lips Threshold: ",touch_lip_threshold)
    print("Face Label: ", findLabel)
    print("Total Frames: ",countLabel[findLabel-1])
    print("Smile Threshold: ",smileDiameterThreshold)
    print("Touch Lips Threshold: ",touch_lip_threshold)
    print("Face Label: ", findLabel)
    print("Total Frames: ",countLabel[findLabel-1])
    print("Lost Frames: ",countLabel[(findLabel)%2])
    print("Eye Calibrated: ",left_cal_h,left_cal_v,right_cal_v,right_cal_h)
    distance_smile =  df[(df.DistanceValue > distanceThreshold) & (df.face_label == findLabel) & (df.SmileDiameter > smileDiameterThreshold) ].count()["Frame"]
    angle_smile =  df[(df.AngleValue < angleThreshold) & (df.face_label == findLabel) & (df.SmileDiameter > smileDiameterThreshold) ].count()["Frame"]
    touch_lips = df[(df.AngleValue < touch_lip_threshold) & (df.face_label == findLabel) & (df.SmileDiameter < 0.8*smileDiameterThreshold) ].count()["Frame"]
    eyeContact = df[(df.Left_V-left_cal_v<2.5) & (df.Left_H-left_cal_h<2.5) & (df.Right_V-right_cal_v<2.5) &(df.Right_H-right_cal_h<2.5) & (df.face_label == findLabel)].count()["Frame"]
    faceTouched = df[(df.FaceTouch == "true") & (df.face_label == findLabel)].count()["Frame"]
    total_frames_count = countLabel[findLabel-1]
    distance_smile_percentage = distance_smile*100/total_frames_count
    angle_smile_percentage = angle_smile*100/total_frames_count
    touch_lips_percentage = touch_lips*100/total_frames_count
    eye_contact_percentage = eyeContact*100/total_frames_count
    face_touch_percentage = faceTouched*100/total_frames_count
    print('Distance Percentage Score: ',distance_smile_percentage)
    print('Angle Percentage Score: ',angle_smile_percentage)
    print('Touch Lips Percentage Score: ',touch_lips_percentage)
    print('Eye Contact Percentage Score: ',eye_contact_percentage)
    print('Face Touch Percentage Score: ',face_touch_percentage)
    return distance_smile_percentage, angle_smile_percentage, touch_lips_percentage,eye_contact_percentage,face_touch_percentage


def run():	
    
    listOfInterviews = INTERVIEW.objects.filter(isVideo = True)
    
    listOfInterviews = listOfInterviews.filter(isVideoProcessed = False)


    metric = THRESHOLD.objects.filter(metricName = "smileLow")[0]
    smileLow = metric.metricThresholdValue
    metric = THRESHOLD.objects.filter(metricName = "smileHigh")[0]
    smileHigh = metric.metricThresholdValue
    metric = THRESHOLD.objects.filter(metricName = "touchLow")[0]
    touchLow = metric.metricThresholdValue
    metric = THRESHOLD.objects.filter(metricName = "touchHigh")[0]
    touchHigh = metric.metricThresholdValue
    metric = THRESHOLD.objects.filter(metricName = "eyeLow")[0]
    eyeLow = metric.metricThresholdValue
    metric = THRESHOLD.objects.filter(metricName = "eyeHigh")[0]
    eyeHigh = metric.metricThresholdValue
    metric = THRESHOLD.objects.filter(metricName = "faceLow")[0]
    faceLow = metric.metricThresholdValue
    metric = THRESHOLD.objects.filter(metricName = "faceHigh")[0]
    faceHigh = metric.metricThresholdValue

    for interview in listOfInterviews:

        print("Interview {0} video processing Start".format(interview.id))
        print(interview)

        smileIndex = None
        touchLipIndex = None
        eyeContactIndex = None
        faceTouchIndex = None

        try:
            st = datetime.now()
            print(interview)
            path = interview.videoFilePath + interview.videoFileName
            distance_sp,angle_sp,touch_sp,eye_sp,face_sp = interviewAnalysis(path)
            ed = datetime.now()
            dif = ed - st
            print("Processing Time: ",dif)
            
            if(distance_sp!="Error"):
                if(distance_sp<=smileLow):
                    smileIndex = "Low"
                if((distance_sp>smileLow) & (distance_sp<smileHigh)):
                    smileIndex = "Medium"
                if(distance_sp>=smileHigh):
                    smileIndex = "High"
                if(touch_sp<=touchLow):
                    touchLipIndex = "Low"
                if((touch_sp>touchLow) & (touch_sp<touchHigh)):
                    touchLipIndex = "Medium"
                if(touch_sp>=touchHigh):
                    touchLipIndex = "High"
                if(eye_sp<=eyeLow):
                    eyeContactIndex = "Low"
                if((eye_sp>eyeLow) & (eye_sp<eyeHigh)):
                    eyeContactIndex = "Medium"
                if(eye_sp>=eyeHigh):
                    eyeContactIndex = "High"
                if(face_sp<=faceLow):
                    faceTouchIndex = "Low"
                if((face_sp>faceLow) & (face_sp<faceHigh)):
                    faceTouchIndex = "Medium"
                if(face_sp>=faceHigh):
                    faceTouchIndex = "High"

        except(Exception):
            print("Error in computing video metrics!!")
            print(traceback.print_exc())

        finally:

            interview.isVideoProcessed = True
            interview.smileIndex = smileIndex
            interview.touchLipIndex = touchLipIndex
            interview.eyeContactIndex = eyeContactIndex
            interview.faceTouchIndex = faceTouchIndex
            
            print("Interview {0} video processed".format(interview.id))
            
            interview.save()
